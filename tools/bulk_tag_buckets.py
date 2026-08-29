"""
bulk_tag_buckets.py — tags videos into curated watchlist buckets so the
library sidebar can group them instead of showing one flat 5,595-item list.

WHAT IT DOES: reads roster sources, matches each video's uploader/channel
name against them, and writes a "bucket" tag onto the match. Each person
lands in exactly one bucket — see EXCLUSIVITY below.

SOURCES (right now):
  - MB_and_Bench_Members.txt   -> "Majority Democrats" and "The Bench" buckets
    (flat — names sit directly inside)
  - house_committee_memberships_119th_current.xlsx (Member_Lookup sheet,
    filtered to Chamber == House) -> "House" bucket (flat, top-level)
  - senate_committee_memberships_119th_current.xlsx (Member_Lookup sheet,
    filtered to Chamber == Senate) -> "Senate" bucket (flat, top-level)
  - federal_cabinet_119th_current.xlsx (Member_Lookup sheet, no Chamber
    filter needed -- every row already IS a sitting Cabinet member) ->
    "Notable Figures" (there's no dedicated Cabinet bucket) (2026-08-27)

This is the same seven-bucket taxonomy archive_items uses (see
app/api/archive/buckets/route.ts): Majority Democrats, The Bench, House,
Senate, Notable Figures, Institutional, Uncategorized — kept in sync
on purpose (2026-08-29) so the two schemas read the same way to a user.

EXCLUSIVITY: each person lands in exactly ONE bucket, picked by priority —
Majority Democrats > The Bench > House > Senate > Notable Figures
(catch-all for anyone matched but not covered by a specific list, including
sitting Cabinet members). A Majority Democrat who's also a sitting House
member shows up under Majority Democrats only, not both. (Note: this is
exclusivity per PERSON, not per VIDEO — a single video whose uploader and
channel fields each resolve to a different real person can still end up
with two bucket tags, one per person. That's expected, not a bug.)

INSTITUTIONAL: a video that matches no roster person at all still lands in
"Institutional" rather than falling straight to Uncategorized if its title
reads as a floor session/hearing/briefing (same title regex
archive_consolidation/enrich_institutional_flag.py uses for the identical
judgment call on archive_items). Anyone matched to neither a person nor
this pattern stays Uncategorized, same as before.

MATCHING: up to four passes per video, in order, stopping at the first hit.

  1. STRICT (name_matches): normalizes names (strips "Rep.", "Sen.", titles,
     periods/commas, extra whitespace, lowercases) and requires the shorter
     name's words to be a full subset of the longer name's words, with at
     least a first+last name overlap — avoids a bare "Josh" accidentally
     matching "Josh Turk" while still tolerating "Rep. Ritchie Torres"
     matching plain "Ritchie Torres".

  2. ALIAS (find_alias_matches, added 2026-08-29): a hand-curated map
     (FIELD_TEXT_ALIASES) of a real, observed nickname/initialism straight
     to the roster entry it means -- e.g. "RFK Jr", which never spells out
     "Kennedy" as a word so passes 1 and 3 can't reach it, and where
     "Kennedy" alone is ambiguous among four different roster members
     anyway (John Kennedy, Mike Kennedy, Timothy M. Kennedy, and Robert F.
     Kennedy, Jr.), so even a spelled-out surname wouldn't resolve to him
     uniquely via pass 3. Checked directly, bypassing ambiguity entirely,
     because it's a hand-verified exact resolution, not an inference.
     Deliberately not a general nickname-guessing system — add one entry
     at a time as a real video surfaces one, same policy as NAME_ALIASES.

  3. SURNAME FALLBACK (added 2026-08-26): campaign/office channels are
     often branded with only a surname — "Rep. Auchincloss", "Mahan for
     California" — which the strict pass can never match: either the
     normalized field collapses to a single word (fails the 2-word-overlap
     floor) or the first name simply never appears in the field text at
     all (fails the subset check). For any video passes 1-2 didn't match,
     this checks whether one of the field's words equals the last word
     (surname) of exactly ONE roster entry. If it's unique, that's a match.
     If more than one roster entry shares that surname (e.g. the known
     "Brendan Boyle" vs. "Brendan F. Boyle" case), it's deliberately left
     unmatched rather than guessing — and logged at the end under
     "ambiguous surnames skipped" so these can be resolved by hand (e.g. by
     adding a distinguishing alias) instead of silently mis-tagging someone.

  4. TITLE FALLBACK (added 2026-08-29): if uploader/channel matched no one
     at all (passes 1-3 above), passes 1-3 run again against the title
     text. Aggregator/news accounts (C-SPAN, a journalist's X/Twitter
     handle, an Instagram news page) routinely carry the actual subject's
     name nowhere but the title -- "President Trump Announces...",
     "Aaron Rupar - RFK Jr complains...".

TAGS: writes TWO tags per match — a kind='bucket' tag (e.g. "Majority
Democrats") for the top-level sidebar grouping, and a kind='person' tag
(e.g. "Ritchie Torres") so the sidebar can nest a subfolder per person
inside each bucket. Both are source='manual', so they never get wiped out
by an auto re-tag (see 0004_tags.sql), and a manual tag intentionally wins
over any auto-generated entity tag with the same name/label.

SAFE TO RE-RUN: uses upsert on (video_id, label), so running it again after
adding Senate/Cabinet/Court sources just adds the new tags without touching
or duplicating what's already there.

DRY RUN BY DEFAULT: matches every video and prints the same counts either
way, but only clears and rewrites tags when --apply is passed. The clear
step also no longer runs until the full new tag set has been computed, so a
crash or network blip while loading videos or matching the roster can't
leave the library's existing bucket/person tags cleared with nothing to
replace them.

Run it from the tools folder, with MB_and_Bench_Members.txt and
house_committee_memberships_119th_current.xlsx in the same folder:

    python bulk_tag_buckets.py             (dry run -- prints counts only)
    python bulk_tag_buckets.py --apply     (actually clear + rewrite tags)
"""

import argparse
import os
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from supabase import create_client, Client

# A Windows console's default codepage (cp1252) can't display every
# character a real video title contains (CJK, emoji, some smart-quote
# variants) -- printing one crashed the whole run partway through the
# ambiguous-surnames summary, which is purely informational, at the exact
# print statement right before the actual --apply tag write. Reconfigure
# rather than let a display-only crash take down the run before it commits.
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# --- CONFIGURATION ---
SUPABASE_URL = os.environ.get("NEXT_PUBLIC_SUPABASE_URL", "https://tijwokimlrglufjqiwok.supabase.co")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
if not SUPABASE_KEY:
    raise SystemExit("SUPABASE_SERVICE_ROLE_KEY must be set in the environment (see .env.local).")

MEMBERS_TXT = Path("MB and Bench Members.txt")
HOUSE_XLSX = Path("house_committee_memberships_119th_current.xlsx")
SENATE_XLSX = Path("senate_committee_memberships_119th_current.xlsx")
CABINET_XLSX = Path("federal_cabinet_119th_current.xlsx")

PAGE_SIZE = 1000

# Every person gets filed under exactly ONE of these, picked by priority —
# not every list they happen to match. Someone who's both a Majority
# Democrat AND a sitting House member goes in Majority Democrats only.
# Anyone matched but not covered by a specific list (including sitting
# Cabinet members -- there's no dedicated Cabinet bucket) falls through to
# the Notable Figures catch-all at the bottom. Mirrors the archive_items
# taxonomy (see app/api/archive/buckets/route.ts) so the two schemas read
# the same way to a user: Majority Democrats, The Bench, House, Senate,
# Notable Figures, Institutional, Uncategorized.
PRIORITY_ORDER = [
    ("Majority Democrats", None),
    ("The Bench", None),
    ("House", None),
    ("Senate", None),
]
CATCH_ALL = ("Notable Figures", None)

# Videos that don't match anyone on the roster at all still deserve a home
# other than plain Uncategorized when they're clearly floor sessions,
# hearings, or briefings rather than something nobody's gotten to yet.
# Same regex archive_consolidation/enrich_institutional_flag.py uses for
# the exact same judgment call on the archive_items side.
INSTITUTIONAL_PATTERNS = re.compile(
    r"(House Session|Senate Session|Morning Hour|Daily Briefing|"
    r"Cabinet Meeting|News Conference|Press Briefing|Speaks to Reporters|"
    r"Republican Agenda|Democratic Agenda|Weekly Briefing|Pen and Pad)",
    re.IGNORECASE,
)

TITLE_PREFIX_RE = re.compile(
    r"^(rep\.?|sen\.?|senator|representative|congressman|congresswoman|gov\.?|governor|mayor)\s+",
    re.IGNORECASE,
)


def normalize_name(name: str) -> str:
    name = (name or "").strip()
    name = TITLE_PREFIX_RE.sub("", name)
    # Punctuation carries no matching-relevant information in a person's
    # name, but its mere PRESENCE broke matching outright: a roster name
    # like "Robert F. Kennedy, Jr." normalized to word set {"f.", "kennedy,",
    # "jr."} (trailing periods/commas glued to the word before the next
    # whitespace split), which real-world uploader/channel/title text almost
    # never reproduces exactly -- "Robert F Kennedy Jr" (no punctuation, the
    # common real-world spelling) failed name_matches's subset check purely
    # because "f" != "f." and "jr" != "jr.", not because the name didn't
    # actually match (2026-08-29).
    name = re.sub(r"[.,]", "", name)
    name = re.sub(r"\s+", " ", name)
    return name.lower().strip()


def name_matches(a: str, b: str) -> bool:
    """a and b are already normalized. Requires the shorter name's words to
    be fully contained in the longer name's words, with a real first+last
    overlap — not just any single shared word."""
    if not a or not b:
        return False
    if a == b:
        return True
    a_words, b_words = set(a.split()), set(b.split())
    shorter, longer = (a_words, b_words) if len(a_words) <= len(b_words) else (b_words, a_words)
    return len(shorter) >= 2 and shorter.issubset(longer)


def build_surname_index(roster: dict) -> dict:
    """Maps a surname (the last word of a normalized roster name) to the
    list of normalized roster names that end in it. A surname mapping to
    more than one roster name is a real ambiguity (e.g. two different
    Boyles) and gets treated as unmatchable by find_surname_fallback —
    never resolved by guessing."""
    idx: dict = defaultdict(list)
    for roster_name in roster:
        words = roster_name.split()
        if not words:
            continue
        idx[words[-1]].append(roster_name)
    return idx


def find_surname_fallback(fields: list, roster: dict, surname_index: dict, ambiguous_log: set) -> dict:
    """Second-pass matcher for videos the strict pass didn't match. Campaign
    and office channels are often branded with only a surname (e.g. "Rep.
    Auchincloss", "Mahan for California") — text the strict two-word-overlap
    rule can never match, since either the field collapses to a single word
    after stripping the title prefix, or the first name never appears in the
    field text at all. This checks whether any word in the field equals the
    surname of exactly one roster entry. Ties (two+ roster entries sharing a
    surname) are recorded in ambiguous_log and deliberately left unmatched."""
    matched = {}
    for field in fields:
        norm_field = normalize_name(field)
        if not norm_field:
            continue
        candidates = set()
        for word in norm_field.split():
            if word in surname_index:
                candidates.update(surname_index[word])
        if len(candidates) == 1:
            roster_name = next(iter(candidates))
            matched[roster_name] = roster[roster_name]
        elif len(candidates) > 1:
            ambiguous_log.add((field, frozenset(candidates)))
    return matched


def parse_member_list(path: Path) -> dict:
    """Parses the '<Bucket Name> (N People)' + blank-line-separated format
    into {bucket_name: [names...]}."""
    header_re = re.compile(r"^(.+?)\s*\(\d+\s*People\)\s*$", re.IGNORECASE)
    sections = {}
    current = None
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line:
            continue
        m = header_re.match(line)
        if m:
            current = m.group(1).strip()
            sections[current] = []
            continue
        if current:
            sections[current].append(line)
    return sections


def load_house_members(path: Path) -> list:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Member_Lookup"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {name: i for i, name in enumerate(header)}
    names = []
    for row in rows:
        if row[idx["Chamber"]] == "House":
            names.append(row[idx["Member Name"]])
    return names


def load_senate_members(path: Path) -> list:
    """Same Member_Lookup shape as load_house_members (2026-08-27, confirmed
    against the actual file: same sheet name, "Member Name"/"Chamber"
    columns present, just Chamber == "Senate" instead of "House") --
    deliberately a separate function rather than a shared chamber="House"/
    "Senate" parameter, so it stays a one-line diff to add if Senate's export
    format ever drifts from House's without disturbing House at all."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Member_Lookup"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {name: i for i, name in enumerate(header)}
    names = []
    for row in rows:
        if row[idx["Chamber"]] == "Senate":
            names.append(row[idx["Member Name"]])
    return names


def load_cabinet_members(path: Path) -> list:
    """A different shape from load_house_members/load_senate_members, not
    just a copy with the filter value swapped: this file's Member_Lookup
    sheet has NO Chamber column, because every row already IS a sitting
    Cabinet member (confirmed 2026-08-27 against the real file: 21 rows,
    Status values only "Confirmed" or "Acting" -- no withdrawn/rejected
    rows to filter out, so no filter is needed at all), and the name column
    here is "Name", not "Member Name" like House/Senate use."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Member_Lookup"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {name: i for i, name in enumerate(header)}
    return [row[idx["Name"]] for row in rows]


# Known name variants that should collapse to ONE person, keyed by their
# normalized (lowercase, title-stripped) form. Deliberately a manual,
# explicit map rather than an automatic rule (e.g. "always strip middle
# initials") -- an automatic rule could just as easily merge two genuinely
# DIFFERENT people who happen to share a first+last name and are only
# distinguished by a middle initial, which is exactly the scenario
# build_surname_index/find_surname_fallback's ambiguous-surname-skip logic
# already exists to protect against elsewhere in this file. Add entries here
# as they're discovered, one line each, rather than guessing a rule.
#
# "Brendan F. Boyle" showed up as a SEPARATE roster entry from "Brendan
# Boyle" -- neither normalize_name nor TITLE_PREFIX_RE strips middle
# initials, so a source that spells him with the "F." produces a second,
# distinct roster key. Any video whose uploader/channel field spelled out
# the full three-word form then satisfied name_matches's subset check
# against BOTH entries at once (the two-word roster name's words are a
# subset of the three-word field, and the three-word field is trivially a
# subset of itself), so that one video got tagged as both people
# simultaneously -- the exact "Brendan Boyle: 276 / Brendan F. Boyle: 1"
# split observed in the library (2026-08-27).
# Note: keys here are post-normalize_name -- normalize_name strips periods/
# commas (2026-08-29), so "Brendan F. Boyle" normalizes to "brendan f boyle",
# not "brendan f. boyle".
NAME_ALIASES: dict[str, str] = {
    "brendan f boyle": "Brendan Boyle",
}

# The President and Vice President aren't covered by ANY of the four
# roster source files above -- MB_and_Bench_Members.txt and the House/
# Senate/Cabinet spreadsheets are all specifically Congress + Cabinet
# rosters, and neither office holder sits in either, so a video whose
# uploader/channel/title names one of them had no roster entry to match
# against at all (confirmed 2026-08-29: neither "Donald Trump" nor "JD
# Vance" existed anywhere in build_roster()'s output). Hand-added here
# rather than parsed from a fifth source file, since these are two people
# who change once every four years at most.
EXECUTIVE_OFFICERS = ["Donald Trump", "JD Vance"]

# Hand-curated map of a real, observed nickname/initialism to the roster
# key it should resolve to -- checked directly, bypassing the surname
# fallback, for names the strict and surname passes genuinely cannot
# reach: "RFK Jr" never spells out "Kennedy" as a word, and "Kennedy"
# alone is ambiguous among four different roster members anyway (John
# Kennedy, Mike Kennedy, Timothy M. Kennedy, and this Robert F. Kennedy,
# Jr.), so even a fully spelled-out surname wouldn't resolve to him
# uniquely. Same "add one entry as it's discovered, never guess a rule"
# policy as NAME_ALIASES above -- this is deliberately not a general
# nickname-guessing system.
FIELD_TEXT_ALIASES: dict[str, str] = {
    "rfk jr": "robert f kennedy jr",
}


def find_alias_matches(fields: list, roster: dict) -> dict:
    """Checked before the surname fallback: does any field contain a
    known alias phrase (FIELD_TEXT_ALIASES)? A direct hit resolves
    straight to that roster entry, sidestepping any surname-ambiguity
    check entirely -- these are hand-verified exact resolutions, not
    inferences that could be wrong."""
    matched = {}
    for field in fields:
        text = normalize_name(field)
        if not text:
            continue
        for alias, roster_name in FIELD_TEXT_ALIASES.items():
            if alias in text and roster_name in roster:
                matched[roster_name] = roster[roster_name]
    return matched


def build_roster() -> dict:
    """Returns {normalized_name: {"display": "Proper Case Name", "memberships": {(bucket, chamber_or_None), ...}}}.
    A person can accumulate several memberships here (e.g. someone who's both
    a Majority Democrat and a sitting House member) — pick_primary_membership()
    is what collapses that down to the one bucket they actually get filed
    under. Every membership here is flat (chamber is always None) -- House
    and Senate are themselves top-level buckets now, not chambers nested
    under a "Federal" parent, and Cabinet has no dedicated bucket at all
    (it's intentionally folded into the Notable Figures catch-all)."""
    roster: dict = {}

    def add_person(display_name: str, bucket_label: str, chamber: str | None = None):
        norm = normalize_name(display_name)
        if not norm:
            return
        canonical = NAME_ALIASES.get(norm)
        if canonical:
            display_name = canonical
            norm = normalize_name(canonical)
        entry = roster.setdefault(norm, {"display": display_name.strip(), "memberships": set()})
        entry["memberships"].add((bucket_label, chamber))

    for bucket_label, names in parse_member_list(MEMBERS_TXT).items():
        for name in names:
            add_person(name, bucket_label)

    for name in load_house_members(HOUSE_XLSX):
        add_person(name, "House")

    for name in load_senate_members(SENATE_XLSX):
        add_person(name, "Senate")

    for name in load_cabinet_members(CABINET_XLSX):
        add_person(name, "Notable Figures")

    for name in EXECUTIVE_OFFICERS:
        add_person(name, "Notable Figures")

    return roster


def pick_primary_membership(memberships: set) -> tuple:
    """A person can technically match several lists (e.g. a Majority
    Democrat who's also a sitting House member). This picks the single
    highest-priority one so they land in exactly one bucket, not several."""
    for candidate in PRIORITY_ORDER:
        if candidate in memberships:
            return candidate
    return CATCH_ALL


def find_matches(fields: list, roster: dict) -> dict:
    """Returns the subset of roster entries (keyed by normalized name) that
    match any of the given video fields — usually zero or one person, but
    a dict in case a video ever legitimately matches more than one."""
    matched = {}
    for field in fields:
        norm_field = normalize_name(field)
        if not norm_field:
            continue
        for roster_name, entry in roster.items():
            if roster_name in matched:
                continue
            if name_matches(norm_field, roster_name):
                matched[roster_name] = entry
    return matched


def fetch_all(supabase: Client, table: str, columns: str) -> list:
    rows, page = [], 0
    while True:
        res = supabase.table(table).select(columns).range(page * PAGE_SIZE, page * PAGE_SIZE + PAGE_SIZE - 1).execute()
        batch = res.data or []
        rows.extend(batch)
        if len(batch) < PAGE_SIZE:
            break
        page += 1
    return rows


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--apply", action="store_true",
                         help="actually clear and rewrite bucket/chamber/person tags (default: dry run, counts only)")
    args = parser.parse_args()

    print("Building roster from local files...")
    roster = build_roster()
    bucket_names = sorted({b for entry in roster.values() for (b, _c) in entry["memberships"]})
    print(f"  {len(roster)} known people across buckets: {', '.join(bucket_names)}")

    surname_index = build_surname_index(roster)
    ambiguous_surnames: set = set()

    print("Connecting to Supabase...")
    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

    print("Loading videos...")
    videos = fetch_all(supabase, "videos", "id, uploader, channel, title")
    print(f"  {len(videos)} videos to check.")

    tag_rows = []
    per_bucket_count = defaultdict(int)
    per_person_count = defaultdict(int)
    matched_videos = 0
    fallback_matched_videos = 0
    title_matched_videos = 0
    institutional_videos = 0

    for v in videos:
        fields = [v.get("uploader"), v.get("channel")]
        matches = find_matches(fields, roster)
        if not matches:
            matches = find_alias_matches(fields, roster)
        if not matches:
            matches = find_surname_fallback(fields, roster, surname_index, ambiguous_surnames)
            if matches:
                fallback_matched_videos += 1
        if not matches:
            # uploader/channel carried no roster person at all -- for a lot of
            # aggregator/news accounts (C-SPAN, Aaron Rupar, Democracy Docket)
            # the subject's name never appears in either field, only in the
            # title itself ("President Trump Announces...", "Rep. Auchincloss
            # holds..."). Some C-SPAN clips report a blank uploader/channel
            # outright, same result. Try the title through the same three
            # passes before giving up on a person match entirely.
            title = v.get("title") or ""
            matches = find_matches([title], roster)
            if not matches:
                matches = find_alias_matches([title], roster)
            if not matches:
                matches = find_surname_fallback([title], roster, surname_index, ambiguous_surnames)
            if matches:
                title_matched_videos += 1
        if not matches:
            # No roster person at all -- still worth filing as Institutional
            # rather than plain Uncategorized if the title itself reads as a
            # floor session/hearing/briefing rather than a person's video.
            if INSTITUTIONAL_PATTERNS.search(v.get("title") or ""):
                institutional_videos += 1
                per_bucket_count["Institutional"] += 1
                tag_rows.append({
                    "video_id": v["id"],
                    "label": "Institutional",
                    "source": "manual",
                    "kind": "bucket",
                })
            continue
        matched_videos += 1
        for entry in matches.values():
            display = entry["display"]
            per_person_count[display] += 1
            tag_rows.append({
                "video_id": v["id"],
                "label": display,
                "source": "manual",
                "kind": "person",
            })
            bucket_label, chamber = pick_primary_membership(entry["memberships"])
            per_bucket_count[bucket_label] += 1
            tag_rows.append({
                "video_id": v["id"],
                "label": bucket_label,
                "source": "manual",
                "kind": "bucket",
            })
            if chamber:
                tag_rows.append({
                    "video_id": v["id"],
                    "label": chamber,
                    "source": "manual",
                    "kind": "chamber",
                })

    print(f"\n{matched_videos} videos matched at least one bucket "
          f"({fallback_matched_videos} of those only via the surname fallback, "
          f"{title_matched_videos} of those only via the title).")
    print(f"{institutional_videos} more videos matched no roster person but read as Institutional by title.")
    for b in sorted(per_bucket_count.keys()):
        print(f"  {b}: {per_bucket_count[b]} videos")
    print(f"\n{len(per_person_count)} distinct people matched at least one video.")

    if ambiguous_surnames:
        print(f"\n{len(ambiguous_surnames)} ambiguous surname(s) skipped (left unmatched rather than guessed):")
        for field, candidates in sorted(ambiguous_surnames, key=lambda x: x[0]):
            print(f"  {field!r} could be: {', '.join(sorted(candidates))}")
        print("  Resolve these by hand (e.g. add a distinguishing alias) if they should match.")

    seen = set()
    deduped_rows = []
    for row in tag_rows:
        key = (row["video_id"], row["label"])
        if key in seen:
            continue
        seen.add(key)
        deduped_rows.append(row)
    tag_rows = deduped_rows

    if not tag_rows:
        print("\nNothing to write — stopping.")
        return

    if not args.apply:
        print(f"\nDry run — {len(tag_rows)} bucket + person tags would be written.")
        print("Re-run with --apply to actually clear the existing bucket/chamber/person "
              "tags and write these.")
        return

    print("\nClearing every bucket/chamber/person tag from previous runs (clean slate)...")
    print("  This only touches kind IN ('bucket','chamber','person') — nothing else.")
    supabase.table("tags").delete().eq("source", "manual").in_("kind", ["bucket", "chamber", "person"]).execute()

    print(f"\nWriting {len(tag_rows)} bucket + person tags...")
    written = 0
    for i in range(0, len(tag_rows), 500):
        chunk = tag_rows[i:i + 500]
        try:
            supabase.table("tags").upsert(chunk, on_conflict="video_id,label").execute()
            written += len(chunk)
        except Exception as e:
            print(f"  FAILED on a batch: {e}")

    print(f"\nDone. {written} bucket tags written (safe to re-run any time).")


if __name__ == "__main__":
    main()
