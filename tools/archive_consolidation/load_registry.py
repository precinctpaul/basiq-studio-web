"""Step 1: load the canonical-ID registry into a fresh working SQLite DB.

Source of truth for this step is CSPAN_YouTube_Master_Registry_v2.xlsx:
  - "Registry"                -> one row per canonical content item
  - "All Files Detail"        -> one row per physical file (matched)
  - "Unmatched (need review)" -> one row per physical file (no canonical ID)

This step does no enrichment and no filesystem access beyond the xlsx
itself -- it just gets the registry's own numbers into a queryable form
so later steps (person/date/quality resolution) have a stable backbone
to attach facts to, and so this step's counts can be checked against the
brief's numbers before anything else is built on top of it.
"""

import argparse

import openpyxl

import config
import schema


def _bool(v) -> int:
    return 1 if v else 0


def load_canonical_items(con, ws) -> int:
    rows = ws.iter_rows(min_row=2, values_only=True)
    n = 0
    for row in rows:
        if row[0] is None:
            continue
        (canonical_id, id_type, project_count, projects, has_video,
         has_raw_video_only, has_transcript, has_metadata,
         has_caption_file, file_count, total_size_mb,
         is_duplicate_across_projects) = row
        con.execute(
            """insert or replace into canonical_items
               (canonical_id, id_type, project_count, projects, has_video,
                has_raw_video_only, has_transcript, has_metadata,
                has_caption_file, file_count, total_size_mb,
                is_duplicate_across_projects)
               values (?,?,?,?,?,?,?,?,?,?,?,?)""",
            (canonical_id, id_type, project_count, projects,
             _bool(has_video), _bool(has_raw_video_only),
             _bool(has_transcript), _bool(has_metadata),
             _bool(has_caption_file), file_count, total_size_mb,
             _bool(is_duplicate_across_projects)),
        )
        n += 1
    return n


def load_files(con, ws) -> int:
    rows = ws.iter_rows(min_row=2, values_only=True)
    n = 0
    for row in rows:
        if row[0] is None:
            continue
        (full_path, name, extension, size_mb, last_write_time, role,
         youtube_id, cspan_id, basiq_uuid, base_folder, project,
         canonical_id, id_type) = row
        con.execute(
            """insert or ignore into files
               (full_path, name, extension, size_mb, last_write_time, role,
                youtube_id, cspan_id, basiq_uuid, base_folder, project,
                canonical_id, id_type)
               values (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (full_path, name, extension, size_mb,
             str(last_write_time) if last_write_time else None, role,
             youtube_id, cspan_id, basiq_uuid, base_folder, project,
             canonical_id, id_type),
        )
        n += 1
    return n


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--fresh", action="store_true",
                         help="delete any existing output DB before loading")
    args = parser.parse_args()

    if args.fresh and config.INDEX_DB.exists():
        config.INDEX_DB.unlink()

    wb = openpyxl.load_workbook(config.REGISTRY_XLSX, read_only=True, data_only=True)
    con = schema.connect(config.INDEX_DB)

    with con:
        n_items = load_canonical_items(con, wb["Registry"])
        # "All Files Detail" already contains every file, including the 133
        # genuinely-unmatched ones (with canonical_id NULL); "Unmatched (need
        # review)" is that same subset filtered for convenience, so it's
        # loaded with INSERT OR IGNORE and contributes no new rows -- it's
        # only here so this loader doesn't silently depend on one sheet
        # agreeing with another without ever checking.
        n_matched = load_files(con, wb["All Files Detail"])
        n_unmatched_sheet = load_files(con, wb["Unmatched (need review)"])

    (n_items_db,) = con.execute("select count(*) from canonical_items").fetchone()
    (n_files_db,) = con.execute("select count(*) from files").fetchone()
    (n_with_video,) = con.execute(
        "select count(*) from canonical_items where has_video = 1"
    ).fetchone()
    (n_dupes,) = con.execute(
        "select count(*) from canonical_items where is_duplicate_across_projects = 1"
    ).fetchone()
    (n_no_canonical,) = con.execute(
        "select count(*) from files where canonical_id is null"
    ).fetchone()
    (n_needs_review,) = con.execute(
        "select count(*) from unmatched_content_files"
    ).fetchone()

    print(f"canonical items loaded: {n_items} (rows in DB: {n_items_db})")
    print(f"matched files loaded:   {n_matched}")
    print(f"unmatched sheet rows loaded (dupes of above, sanity check): {n_unmatched_sheet}")
    print(f"total file rows in DB:  {n_files_db}")
    print(f"canonical items with video: {n_with_video}")
    print(f"canonical items duplicated across projects: {n_dupes}")
    print(f"files with no canonical_id at all (scripts/logs/db files included): {n_no_canonical}")
    print(f"files genuinely needing review (video/metadata_json/transcript_csv "
          f"with no canonical_id): {n_needs_review}  <- should be 133 per the brief")
    print(f"\nwrote {config.INDEX_DB}")

    con.close()


if __name__ == "__main__":
    main()
