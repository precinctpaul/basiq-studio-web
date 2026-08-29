"""BioGuideID reference data: the full historical roster, and the current
119th-Congress subset used to scope fuzzy name matching.

Fuzzy-matching a name found in a title/description against all 12,769
historical members risks false hits from decades-old name collisions
(there have been multiple "Murray"s, "Smith"s, etc. across 230+ years of
Congress). Every item in this archive is 2025-2026 footage, so matching
is restricted to sitting 119th Congress members -- confirmed against
tools/us_political_directory_master.xlsx (generated 2026-08-27, i.e.
current as of this build), not guessed from birth year or any other
proxy. Historical footage, if any turns out to exist, is expected to fall
through to Unmatched-No-Person rather than risk a wrong match.
"""

import csv
from pathlib import Path

import openpyxl

import config

POLITICAL_DIRECTORY_XLSX = config.TOOL_DIR.parent / "us_political_directory_master.xlsx"


def load_bioguide_reference() -> dict[str, dict]:
    """All 12,769 historical members, keyed by bioguide_id."""
    with open(config.BIOGUIDE_CSV, encoding="utf-8") as f:
        return {row["bioguide_id"]: row for row in csv.DictReader(f)}


def load_current_bioguide_ids() -> set[str]:
    """Every bioguide_id appearing in the 119th Congress committee rosters
    (House + Senate), deduplicated across committee assignments."""
    wb = openpyxl.load_workbook(POLITICAL_DIRECTORY_XLSX, read_only=True, data_only=True)
    ids: set[str] = set()
    for sheet_name in ("House_Committees", "Senate_Committees"):
        ws = wb[sheet_name]
        header = next(ws.iter_rows(max_row=1, values_only=True))
        bioguide_col = header.index("Bioguide ID")
        for row in ws.iter_rows(min_row=2, values_only=True):
            bid = row[bioguide_col]
            if bid:
                ids.add(bid)
    return ids


def build_current_member_index() -> dict[str, list[str]]:
    """Map a lowercase 'first last' name string to the bioguide_id(s) that
    could match it, restricted to current members. A list (not a single
    id) because two sitting members can share a first+last name."""
    reference = load_bioguide_reference()
    current_ids = load_current_bioguide_ids()

    index: dict[str, list[str]] = {}
    missing_from_reference = []
    for bid in current_ids:
        row = reference.get(bid)
        if not row:
            missing_from_reference.append(bid)
            continue
        key = f"{row['first_name']} {row['last_name']}".strip().lower()
        index.setdefault(key, []).append(bid)

    if missing_from_reference:
        print(f"warning: {len(missing_from_reference)} current bioguide_ids from the "
              f"committee rosters are not in the reference CSV: {missing_from_reference[:10]}")

    return index


if __name__ == "__main__":
    current_ids = load_current_bioguide_ids()
    print(f"current 119th Congress bioguide_ids (House+Senate committees): {len(current_ids)}")
    idx = build_current_member_index()
    print(f"distinct 'first last' name keys: {len(idx)}")
    collisions = {k: v for k, v in idx.items() if len(v) > 1}
    print(f"name collisions among current members: {len(collisions)}")
    for k, v in collisions.items():
        print(f"  {k}: {v}")
